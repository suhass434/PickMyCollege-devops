import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pymongo import MongoClient
import warnings
import argparse

warnings.filterwarnings('ignore')

# Configuration
COLLEGE_LOCATION_CSV = 'shared/college_location.csv'
BRANCH_CODE_CSV = 'shared/branch_code.csv'
OUTPUT_DIR = 'extractor/output/'
OUTPUT_XLSX = 'extractor/output/engg_cutoff_gen_from_csv2.xslx'
MONGO_URI = "mongodb+srv://pickyourcollege-admin:xsVTHOPFiktTloKM@pickmycollege.slpeom5.mongodb.net/?retryWrites=true&w=majority&appName=PickMyCollege"

def read_college_locations():
    """Read college code to name and location mapping"""
    df = pd.read_csv(COLLEGE_LOCATION_CSV)
    # Return a dictionary with both name and location
    return {
        row['college_code'].strip().upper(): {
            'name': row['college_name'].strip(),
            'location': row['location'].strip()
        }
        for _, row in df.iterrows()
    }

def read_branch_codes():
    """Read branch code to branch name mapping"""
    df = pd.read_csv(BRANCH_CODE_CSV)
    # Assuming the CSV has columns like 'branch_code' and 'branch_name'
    return {
        row['code'].strip().upper(): row['name'].strip()
        for _, row in df.iterrows()
    }
    
def save_to_mongodb(records, year):
    """Save records to MongoDB (from extract.py)"""
    client = MongoClient(MONGO_URI)
    db = client["cet"]
    collection = db[year]
    
    if records:
        collection.insert_many(records)
        print(f"[✓] Inserted {len(records)} documents into MongoDB → cet.{year}")
    else:
        print("[!] No records to insert into MongoDB.")

def process_excel(INPUT_XLSX):
    """Main processing function for XLSX input"""
    # Load college locations and names
    college_data = read_college_locations()

    # Load branch codes and names
    branch_data = read_branch_codes()

    # Read Excel file
    df = pd.read_excel(INPUT_XLSX, header=None)
    records = []
    current_college = None
    categories = []
    processing_table = False

    for index, row in df.iterrows():
        # Skip empty rows
        if row.isnull().all():
            continue
        
        # Convert row to list of strings
        row_data = [str(cell).strip() if pd.notnull(cell) else '' for cell in row]
        
        # Detect college code (E followed by numbers)
        college_match = re.search(r'(E\d+)\s+(.*)', ' '.join(row_data), re.IGNORECASE)
        if college_match:
            # Handle back-to-back college entries
            if current_college and not processing_table:
                print(f"Warning: Overwriting previous college {current_college['code']} with new entry")
                
            code = college_match.groups()[0].upper().strip()

            # Get college name and location from CSV instead of Excel
            if code in college_data:
                current_college = {
                    'code': code,
                    'name': college_data[code]['name'],
                    'location': college_data[code]['location']
                }
            else:
                # Fallback to Excel data if not found in CSV
                name = college_match.groups()[1].strip()
                current_college = {
                    'code': code,
                    'name': name,
                    'location': ''
                }
                print(f"Warning: College code {code} not found in {COLLEGE_LOCATION_CSV}, using Excel data")
                print(f"{code} - {name}")
                
            processing_table = False
            continue
        
        # Detect table headers (categories) using regex
        if current_college and not processing_table:
            # Check if at least 3 cells match the category pattern (e.g., 1G, 2AG)
            category_candidates = [re.match(r'^\d+[A-Za-z]+$', cell.strip()) for cell in row_data]
            if sum(bool(m) for m in category_candidates) >= 3:
                categories = [cell.upper().strip() for cell in row_data]
                category_col_indexes = {idx: cat for idx, cat in enumerate(categories) if cat}
                processing_table = True
                continue

        # Process table rows
        if current_college and processing_table:
            if not any(cell for cell in row_data):
                continue

            # Extract branch code and name using regex
            branch_info = row_data[0].split(' ', 1)
            branch_code = branch_info[0] if len(branch_info) > 1 else ''
            # Get branch name from CSV instead of Excel
            if branch_code.upper() in branch_data:
                branch_name = branch_data[branch_code.upper()]
            else:
                # Fallback to Excel data if not found in CSV
                branch_name = branch_info[1] if len(branch_info) > 1 else row_data[0]
                if branch_code:  # Only show warning if there was a branch code
                    print(f"Warning: Branch code {branch_code} not found in {BRANCH_CODE_CSV}, using Excel data.")
                    print(f"{branch_code} - {branch_name}")

            # Process cutoff values using explicit category indexes
            for col_idx, category in category_col_indexes.items():
                if col_idx >= len(row_data):
                    continue
                value = row_data[col_idx].replace(',', '').strip()
                if value in ('', '--', 'nan'):
                    continue

                try:
                    cutoff = int(float(value))
                    if 1 <= cutoff <= 500000:  # Validate realistic cutoff range
                        records.append({
                            'college_code': current_college['code'],
                            'college_name': current_college['name'],
                            'location': current_college['location'],
                            'branch_code': branch_code,
                            'branch_name': branch_name,
                            'category': category,
                            'cutoff': cutoff
                        })
                except (ValueError, TypeError):
                    continue

    return records

def export_to_excel(records):
    """Export records to Excel with formatting"""
    df = pd.DataFrame(records)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = ['college_code', 'college_name', 'location', 
               'branch_code', 'branch_name', 'category', 'cutoff']
    ws.append(headers)
    
    # Add data and apply formatting
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
        
        # Highlight cutoff values above 2 lakh
        cutoff_cell = ws.cell(row=ws.max_row, column=headers.index('cutoff')+1)
        if cutoff_cell.value and cutoff_cell.value > 200000:
            cutoff_cell.fill = red_fill
            
    # Save file
    wb.save(OUTPUT_XLSX)
    
def export_to_csv(records, year):
    """Export records to CSV"""
    df = pd.DataFrame(records)
    
    # Define column order
    columns = ['college_code', 'college_name', 'location', 
               'branch_code', 'branch_name', 'category', 'cutoff']
    
    # Reorder columns and save to CSV
    df = df[columns]
    df.to_csv(f"{OUTPUT_DIR}{year}.csv", index=False)

def main(INPUT_XLSX):
    # Process Excel file
    records = process_excel(INPUT_XLSX)
    
    year = input("Enter year (e.g., 2024): ").strip()

    # Export to Excel
    export_to_csv(records, year)
    print(f"Exported {len(records)} records to {OUTPUT_DIR}{year}.csv")
    
    # Save to MongoDB
    save_to_mongodb(records, year)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Extract college cutoff data from Excel file')
    parser.add_argument('input_file', help='Path to the input Excel file')
    
    args = parser.parse_args()
    main(args.input_file)